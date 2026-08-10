"""Export stored PubMed articles to Excel for offline / LLM analysis."""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Any, Sequence

from ep_monitor import config

logger = logging.getLogger(__name__)

_COLUMNS = [
    "pmid",
    "title",
    "abstract",
    "journal",
    "publication_date",
    "authors",
    "url",
    "matched_companies",
    "matched_products",
    "domain_id",
    "fetched_at",
    "source",
]


def export_articles_to_excel(
    rows: Sequence[dict[str, Any]],
    *,
    output_dir: Path | None = None,
    report_date: date | None = None,
    filename: str | None = None,
) -> Path:
    """Write article rows to ``.xlsx``. Returns the output path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel export. Run: pip install -r requirements.txt"
        ) from exc

    out_dir = output_dir or config.EXPORTS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    day = (report_date or date.today()).isoformat()
    path = out_dir / (filename or f"articles_{day}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"

    header_fill = PatternFill("solid", fgColor="C8102E")
    header_font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = header_fill
        cell.font = header_font

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(_COLUMNS, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value if value is not None else "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 12,
        "B": 48,
        "C": 60,
        "D": 24,
        "E": 14,
        "F": 28,
        "G": 28,
        "H": 22,
        "I": 22,
        "J": 12,
        "K": 20,
        "L": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(path)
    logger.info("Exported %d article(s) to %s", len(rows), path)
    return path


def articles_to_rows(articles: Sequence[Any], *, domain_id: str = "") -> list[dict[str, Any]]:
    """Convert ``Article`` objects to Excel-ready dict rows."""
    rows: list[dict[str, Any]] = []
    for article in articles:
        rows.append(
            {
                "pmid": article.source_id,
                "title": article.title,
                "abstract": article.abstract or "",
                "journal": article.journal or "",
                "publication_date": (
                    article.publication_date.isoformat() if article.publication_date else ""
                ),
                "authors": ", ".join(article.authors or []),
                "url": article.url or "",
                "matched_companies": ", ".join(article.matched_companies or []),
                "matched_products": ", ".join(article.matched_products or []),
                "domain_id": domain_id,
                "fetched_at": "",
                "source": article.source,
            }
        )
    return rows
