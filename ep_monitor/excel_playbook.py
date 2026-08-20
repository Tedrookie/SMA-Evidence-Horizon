"""Excel playbook template + import/export for Evidence Horizon.

Medical Affairs can edit domains, keywords, and competitor company names
in Excel, then upload the file in the App Manual console.
Competitor product lists are optional.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, BinaryIO

from ep_monitor import config
from ep_monitor import playbook as pb

logger = logging.getLogger(__name__)

TEMPLATE_PATH = config.DATA_DIR / "evidence_horizon_playbook_template.xlsx"

_JJ_RED = "C8102E"


def write_playbook_excel(playbook: dict[str, Any] | None = None, path: Path | None = None) -> Path:
    """Write a playbook workbook MA can edit and re-upload."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    book = playbook if playbook is not None else pb.load_playbook()
    out = path or TEMPLATE_PATH
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor=_JJ_RED)
    header_font = Font(color="FFFFFF", bold=True)

    # --- Settings ---
    ws = wb.active
    ws.title = "Settings"
    ws.append(["key", "value", "notes"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    meta = book.get("meta") or {}
    sched = book.get("schedule") or {}
    rows = [
        ("product_name", meta.get("product_name", "Evidence Horizon"), "System display name"),
        ("tagline", meta.get("tagline", ""), "Slogan at the top of the email"),
        (
            "digest_title",
            meta.get("digest_title", ""),
            "Issue title after SMA Evidence Horizon: (e.g. Stenosis in Neurovascular)",
        ),
        ("owner", meta.get("owner", "Strategic Medical Affairs / JJMC"), ""),
        ("schedule_mode", sched.get("mode", "weekly"), "weekly or daily"),
        ("weekday", sched.get("weekday", "monday"), "For weekly runs"),
        ("hour", sched.get("hour", 8), "Local hour 0-23"),
        ("minute", sched.get("minute", 0), "Local minute 0-59"),
        ("lookback_days", sched.get("lookback_days", 7), "PubMed lookback window"),
        (
            "recipients",
            "; ".join(
                (r.get("email") if isinstance(r, dict) else str(r))
                for r in (book.get("recipients") or [])
                if (isinstance(r, dict) and r.get("email")) or (not isinstance(r, dict) and str(r).strip())
            ),
            "Emails separated by ;  (BCC)",
        ),
    ]
    for row in rows:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 36

    # --- Domains ---
    wd = wb.create_sheet("Domains")
    wd.append(
        [
            "domain_id",
            "domain_name",
            "enabled",
            "technologies",
            "diseases",
            "topics",
        ]
    )
    for cell in wd[1]:
        cell.fill = header_fill
        cell.font = header_font
    for domain in book.get("domains") or []:
        if not isinstance(domain, dict):
            continue
        wd.append(
            [
                domain.get("id", ""),
                domain.get("name", ""),
                "TRUE" if domain.get("enabled", True) else "FALSE",
                " | ".join(domain.get("technologies") or []),
                " | ".join(domain.get("diseases") or []),
                " | ".join(domain.get("ep_topics") or []),
            ]
        )
    for col, width in {"A": 14, "B": 28, "C": 10, "D": 45, "E": 40, "F": 40}.items():
        wd.column_dimensions[col].width = width

    # --- Companies ---
    wc = wb.create_sheet("Companies")
    wc.append(["domain_id", "company_name", "role", "products_optional"])
    for cell in wc[1]:
        cell.fill = header_fill
        cell.font = header_font
    for domain in book.get("domains") or []:
        if not isinstance(domain, dict):
            continue
        did = domain.get("id", "")
        for row in domain.get("companies") or []:
            if not isinstance(row, dict):
                continue
            wc.append(
                [
                    did,
                    row.get("name", ""),
                    "competitor",
                    " | ".join(str(p) for p in (row.get("products") or [])),
                ]
            )
        for row in domain.get("own_portfolio") or []:
            if not isinstance(row, dict):
                continue
            wc.append(
                [
                    did,
                    row.get("name", ""),
                    "jj",
                    " | ".join(str(p) for p in (row.get("products") or [])),
                ]
            )
    for col, width in {"A": 14, "B": 28, "C": 12, "D": 40}.items():
        wc.column_dimensions[col].width = width

    # --- Instructions ---
    wi = wb.create_sheet("Instructions", 0)
    wi["A1"] = "Evidence Horizon — Playbook Excel (for Medical Affairs)"
    wi["A1"].font = Font(bold=True, size=14, color=_JJ_RED)
    instructions = [
        "",
        "How to use",
        "1. Edit Settings / Domains / Companies sheets (do not rename sheets or header row).",
        "2. In Settings, digest_title appears in email as: SMA Evidence Horizon: <your phrase>.",
        "3. In Domains, separate multiple keywords with  |  (space-pipe-space).",
        "4. In Companies, role must be competitor or jj (Johnson & Johnson portfolio).",
        "5. products_optional can be left blank — company name alone is enough for search.",
        "6. Add a new domain: new row in Domains + company rows with the same domain_id.",
        "7. Upload this file in the Evidence Horizon console (Playbook Excel upload).",
        "",
        "Starter domains: ep (Electrophysiology), nv (Neurovascular), surgery (Surgery).",
    ]
    for i, line in enumerate(instructions, start=2):
        wi[f"A{i}"] = line
    wi.column_dimensions["A"].width = 100

    wb.save(out)
    logger.info("Wrote playbook Excel to %s", out)
    return out


def _split_keywords(raw: Any) -> list[str]:
    text = str(raw or "").strip()
    if not text:
        return []
    parts = [p.strip() for p in text.replace(";", "|").split("|")]
    return [p for p in parts if p]


def _as_bool(raw: Any) -> bool:
    text = str(raw or "").strip().lower()
    return text in {"1", "true", "yes", "y", "enabled"}


def playbook_from_excel(file: Path | BinaryIO | bytes) -> dict[str, Any]:
    """Parse an uploaded/edited playbook workbook into a playbook dict."""
    from openpyxl import load_workbook

    if isinstance(file, (bytes, bytearray)):
        from io import BytesIO

        wb = load_workbook(BytesIO(file), data_only=True)
    else:
        wb = load_workbook(file, data_only=True)

    base = pb.default_playbook()
    meta = dict(base.get("meta") or {})
    sched = dict(base.get("schedule") or {})
    recipients: list[dict[str, str]] = []

    if "Settings" in wb.sheetnames:
        ws = wb["Settings"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            key = str(row[0]).strip()
            value = row[1]
            if key == "product_name":
                meta["product_name"] = str(value or meta.get("product_name", "")).strip()
            elif key == "tagline":
                meta["tagline"] = str(value or "").strip()
            elif key == "digest_title":
                meta["digest_title"] = str(value or "").strip()
            elif key == "owner":
                meta["owner"] = str(value or "").strip()
            elif key == "schedule_mode":
                sched["mode"] = str(value or "weekly").strip().lower()
            elif key == "weekday":
                sched["weekday"] = str(value or "monday").strip().lower()
            elif key == "hour":
                try:
                    sched["hour"] = int(value)
                except (TypeError, ValueError):
                    pass
            elif key == "minute":
                try:
                    sched["minute"] = int(value)
                except (TypeError, ValueError):
                    pass
            elif key == "lookback_days":
                try:
                    sched["lookback_days"] = int(value)
                except (TypeError, ValueError):
                    pass
            elif key == "recipients":
                for part in str(value or "").replace(",", ";").split(";"):
                    email = part.strip()
                    if email:
                        recipients.append({"name": "", "email": email})

    domains_by_id: dict[str, dict[str, Any]] = {}
    if "Domains" in wb.sheetnames:
        ws = wb["Domains"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            did = str(row[0]).strip()
            domains_by_id[did] = {
                "id": did,
                "name": str(row[1] or did).strip(),
                "enabled": _as_bool(row[2]) if row[2] is not None else True,
                "technologies": _split_keywords(row[3] if len(row) > 3 else ""),
                "diseases": _split_keywords(row[4] if len(row) > 4 else ""),
                "ep_topics": _split_keywords(row[5] if len(row) > 5 else ""),
                "companies": [],
                "own_portfolio": [],
            }

    if "Companies" in wb.sheetnames:
        ws = wb["Companies"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not row[1]:
                continue
            did = str(row[0]).strip()
            name = str(row[1]).strip()
            role = str(row[2] or "competitor").strip().lower()
            products = _split_keywords(row[3] if len(row) > 3 else "")
            if did not in domains_by_id:
                domains_by_id[did] = {
                    "id": did,
                    "name": did,
                    "enabled": True,
                    "technologies": [],
                    "diseases": [],
                    "ep_topics": [],
                    "companies": [],
                    "own_portfolio": [],
                }
            entry = {"name": name, "products": products}
            if role in {"jj", "j&j", "jnj", "own", "own_portfolio"}:
                domains_by_id[did]["own_portfolio"].append(entry)
            else:
                domains_by_id[did]["companies"].append(entry)

    domains = list(domains_by_id.values())
    if not domains:
        raise ValueError("No domains found in Excel. Check the Domains sheet.")

    return {
        "meta": meta,
        "schedule": sched,
        "recipients": recipients,
        "domains": domains,
    }


def ensure_template() -> Path:
    """Create the starter Excel template if missing."""
    if not TEMPLATE_PATH.exists():
        return write_playbook_excel(pb.load_playbook(), TEMPLATE_PATH)
    return TEMPLATE_PATH
